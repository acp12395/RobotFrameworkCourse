import openpyxl
from robot.api.deco import keyword, library

@library
class BookHandler:
    def __init__(self):
        self._book = None
        self._sheet = None
    @keyword
    def open_file_sheet(self, file_name, sheet_name):
        self._book = openpyxl.load_workbook(file_name)
        self._sheet = self._book[sheet_name]
    @keyword
    def get_number_of_rows(self):
        return self._sheet.max_row
    @keyword
    def get_number_of_columns(self):
        return self._sheet.max_column
    @keyword
    def get_cell_value(self, row, column):
        return self._sheet.cell(row, column).value

"""obj = BookHandler()
obj.open_file_sheet("Book.xlsx","Sheet1")
cols = obj.get_number_of_columns()
print(cols)
rows = obj.get_number_of_rows()
print(rows)
for row in range(2, rows+1):
    for col in range(1, cols+1):
        field = obj.get_cell_value(1, col)
        value = obj.get_cell_value(row, col)
        if value == None:
            continue
        print(field, value)"""